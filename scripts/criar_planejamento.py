# scripts/criar_planejamento.py
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DISCIPLINAS = ["Português", "Matemática", "Ciências", "História", "Geografia", "Artes"]
BIMESTRES   = ["B1", "B2", "B3", "B4"]

COR_AZUL   = "1F4E79"
COR_BRANCO = "FFFFFF"

def _fill(cor):
    return PatternFill(fill_type="solid", fgColor=cor)

def _font(bold=False, cor=COR_BRANCO, size=10):
    return Font(bold=bold, color=cor, size=size)

def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def aplicar_header(cell, valor, cor=COR_AZUL):
    cell.value = valor
    cell.font      = _font(bold=True)
    cell.fill      = _fill(cor)
    cell.alignment = _center(wrap=True)
    cell.border    = _border()

def add_calendario(wb):
    ws = wb.create_sheet("Calendário Letivo")
    ws.sheet_properties.tabColor = "1F4E79"

    # Seção: Bimestres
    ws.merge_cells("A1:D1")
    t = ws.cell(row=1, column=1, value="CALENDÁRIO LETIVO — 1º ANO")
    t.font = Font(bold=True, size=13, color=COR_BRANCO)
    t.fill = _fill(COR_AZUL)
    t.alignment = _center()
    ws.row_dimensions[1].height = 30

    bim_headers = ["Bimestre", "Início", "Término", "Obs."]
    bim_widths  = [16, 14, 14, 40]
    for col, (h, w) in enumerate(zip(bim_headers, bim_widths), 1):
        aplicar_header(ws.cell(row=2, column=col), h)
        ws.column_dimensions[get_column_letter(col)].width = w

    bimestres_datas = [
        ("1º Bimestre", "03/02/2025", "25/04/2025", ""),
        ("2º Bimestre", "28/04/2025", "04/07/2025", ""),
        ("3º Bimestre", "28/07/2025", "26/09/2025", ""),
        ("4º Bimestre", "29/09/2025", "05/12/2025", ""),
    ]
    for r, (b, ini, fim, obs) in enumerate(bimestres_datas, 3):
        for col, val in enumerate([b, ini, fim, obs], 1):
            c = ws.cell(row=r, column=col, value=val)
            c.border    = _border()
            c.alignment = _center() if col < 4 else Alignment(horizontal="left", vertical="center")

    # Seção: Datas importantes
    ws.cell(row=8, column=1, value="DATAS IMPORTANTES").font = Font(bold=True, size=11)
    ws.row_dimensions[8].height = 25

    di_headers = ["Data", "Evento", "Tipo"]
    di_widths  = [14, 50, 25]
    for col, (h, w) in enumerate(zip(di_headers, di_widths), 1):
        aplicar_header(ws.cell(row=9, column=col), h)
        ws.column_dimensions[get_column_letter(col)].width = w

    eventos = [
        ("25/01/2025", "Início do ano letivo", "Institucional"),
        ("01/05/2025", "Dia do Trabalho — Feriado", "Feriado"),
        ("12/06/2025", "Corpus Christi", "Feriado"),
        ("25/07/2025", "Retorno 2º semestre", "Institucional"),
        ("07/09/2025", "Independência — Feriado", "Feriado"),
        ("12/10/2025", "N. Sra. Aparecida — Feriado", "Feriado"),
        ("02/11/2025", "Finados — Feriado", "Feriado"),
        ("15/11/2025", "Proclamação da República — Feriado", "Feriado"),
    ]
    for r, (data, evento, tipo) in enumerate(eventos, 10):
        for col, val in enumerate([data, evento, tipo], 1):
            c = ws.cell(row=r, column=col, value=val)
            c.border    = _border()
            c.alignment = _center() if col != 2 else Alignment(horizontal="left", vertical="center")

    # Linhas em branco para a professora adicionar feriados municipais
    for r in range(18, 30):
        for col in range(1, 4):
            ws.cell(row=r, column=col).border = _border()

    ws.freeze_panes = "A3"


def add_planejamento_bimestre(wb, bimestre, n_semanas):
    ws = wb.create_sheet(f"{bimestre} - Planejamento")
    ws.sheet_properties.tabColor = "375623"

    # Linha 1: título
    n_cols = 2 + len(DISCIPLINAS) * 4
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    t = ws.cell(row=1, column=1, value=f"PLANEJAMENTO SEMANAL — {bimestre}")
    t.font = Font(bold=True, size=12, color=COR_BRANCO)
    t.fill = _fill(COR_AZUL)
    t.alignment = _center()
    ws.row_dimensions[1].height = 30

    # Linha 2: cabeçalhos fixos
    aplicar_header(ws.cell(row=2, column=1), "Semana")
    aplicar_header(ws.cell(row=2, column=2), "Período")
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 22

    # Cabeçalhos de disciplinas (cada disciplina ocupa 4 colunas)
    sub_headers = ["Conteúdo / Tema", "Objetivo", "Atividade / Recurso", "Realizado?"]
    sub_widths  = [30, 25, 25, 11]
    COR_DISC = ["2E75B6", "375623", "7030A0", "C55A11", "C00000", "4472C4"]

    for d_idx, disc in enumerate(DISCIPLINAS):
        col_base = 3 + d_idx * 4
        # Merge para nome da disciplina
        ws.merge_cells(
            start_row=2, start_column=col_base,
            end_row=2,   end_column=col_base + 3
        )
        aplicar_header(ws.cell(row=2, column=col_base), disc, cor=COR_DISC[d_idx])

        # Sub-headers na linha 3
        for s_idx, (sh, sw) in enumerate(zip(sub_headers, sub_widths)):
            aplicar_header(ws.cell(row=3, column=col_base + s_idx), sh, cor=COR_DISC[d_idx])
            ws.column_dimensions[get_column_letter(col_base + s_idx)].width = sw

    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 28

    # Linhas de semanas
    for sem in range(1, n_semanas + 1):
        row = 3 + sem
        ws.cell(row=row, column=1, value=f"Semana {sem}")
        ws.cell(row=row, column=2, value="")  # professora preenche o período

        for col in range(1, n_cols + 1):
            c = ws.cell(row=row, column=col)
            c.border = _border()
            is_realizado = (col - 3) % 4 == 3 and col >= 3
            c.alignment = Alignment(
                horizontal="center" if col == 1 or is_realizado else "left",
                vertical="top",
                wrap_text=True
            )
        ws.row_dimensions[row].height = 60

    # Semana de recuperação (última linha, destaque)
    row_recup = 3 + n_semanas + 1
    ws.merge_cells(f"A{row_recup}:{get_column_letter(n_cols)}{row_recup}")
    c = ws.cell(row=row_recup, column=1, value="SEMANA DE RECUPERAÇÃO / AVALIAÇÃO")
    c.font = Font(bold=True, color=COR_BRANCO)
    c.fill = _fill("FF0000")
    c.alignment = _center()
    c.border = _border()
    ws.row_dimensions[row_recup].height = 25

    ws.freeze_panes = "C4"


def main():
    saida = os.path.join(os.path.dirname(__file__), "..", "planejamento-semanal.xlsx")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    add_calendario(wb)
    # Semanas por bimestre (ajuste conforme o calendário da escola)
    semanas_por_bimestre = {"B1": 10, "B2": 10, "B3": 10, "B4": 9}
    for b, n in semanas_por_bimestre.items():
        add_planejamento_bimestre(wb, b, n)

    wb.save(saida)
    print(f"Gerado: {os.path.abspath(saida)}")

if __name__ == "__main__":
    main()
