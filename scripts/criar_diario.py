# scripts/criar_diario.py
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

N_ALUNOS   = 27
DISCIPLINAS = ["Português", "Matemática", "Ciências", "História", "Geografia", "Artes"]
BIMESTRES  = ["B1", "B2", "B3", "B4"]
MESES      = ["Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov"]

COR_AZUL   = "1F4E79"
COR_BRANCO = "FFFFFF"
COR_ROXO   = "D9D2E9"   # destaque aluno especial

def _fill(cor):
    return PatternFill(fill_type="solid", fgColor=cor)

def _font(bold=False, cor=COR_BRANCO, size=10):
    return Font(bold=bold, color=cor, size=size)

def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def aplicar_header(cell, valor):
    cell.value = valor
    cell.font   = _font(bold=True)
    cell.fill   = _fill(COR_AZUL)
    cell.alignment = _center(wrap=True)
    cell.border = _border()

def estilizar_dado(cell, centralizar=True):
    cell.border = _border()
    if centralizar:
        cell.alignment = _center()

def add_turma(wb):
    ws = wb.create_sheet("Turma")
    ws.sheet_properties.tabColor = "1F4E79"

    colunas = [
        ("Nº",               5),
        ("Nome Completo",   35),
        ("Dt. Nascimento",  16),
        ("Responsável",     30),
        ("Contato",         18),
        ("Al. Especial",    13),
        ("Observações",     40),
    ]
    for col, (titulo, largura) in enumerate(colunas, 1):
        aplicar_header(ws.cell(row=1, column=col), titulo)
        ws.column_dimensions[get_column_letter(col)].width = largura

    ws.row_dimensions[1].height = 30

    for row in range(2, N_ALUNOS + 2):
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=6, value="Não")
        for col in range(1, len(colunas) + 1):
            estilizar_dado(ws.cell(row=row, column=col))

    ws.freeze_panes = "B2"
    return ws

def add_presenca_sheets(wb):
    red_fill = _fill("FFC7CE")

    for mes in MESES:
        ws = wb.create_sheet(f"Presença - {mes}")
        ws.sheet_properties.tabColor = "2E75B6"

        # Cabeçalhos fixos
        aplicar_header(ws.cell(row=1, column=1), "Nº")
        aplicar_header(ws.cell(row=1, column=2), "Nome")

        # Dias 1-31
        for dia in range(1, 32):
            aplicar_header(ws.cell(row=1, column=2 + dia), str(dia))
            ws.column_dimensions[get_column_letter(2 + dia)].width = 4

        # Totalizadores
        for titulo, col in [("Total P", 34), ("Total F", 35), ("Total J", 36), ("% Freq", 37)]:
            aplicar_header(ws.cell(row=1, column=col), titulo)
            ws.column_dimensions[get_column_letter(col)].width = 9

        ws.row_dimensions[1].height = 25
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 30

        for row in range(2, N_ALUNOS + 2):
            ws.cell(row=row, column=1, value=row - 1)
            ws.cell(row=row, column=2).value = f"=Turma!B{row}"

            # Fórmulas de contagem: dias ficam em C-AG (cols 3-33)
            ws.cell(row=row, column=34).value = f'=COUNTIF(C{row}:AG{row},"P")'
            ws.cell(row=row, column=35).value = f'=COUNTIF(C{row}:AG{row},"F")'
            ws.cell(row=row, column=36).value = f'=COUNTIF(C{row}:AG{row},"J")'
            ws.cell(row=row, column=37).value = (
                f"=IF((AH{row}+AI{row}+AJ{row})=0,"
                f'"-",AH{row}/(AH{row}+AI{row}+AJ{row}))'
            )
            ws.cell(row=row, column=37).number_format = "0.0%"

            for col in range(1, 38):
                estilizar_dado(ws.cell(row=row, column=col))

        # Alerta vermelho para % Freq < 75%
        ws.conditional_formatting.add(
            f"AK2:AK{N_ALUNOS + 1}",
            CellIsRule(operator="lessThan", formula=["0.75"], fill=red_fill)
        )

        ws.freeze_panes = "C2"

def add_notas_sheets(wb):
    red_fill  = _fill("FFC7CE")   # Reprovado: nota < 5
    blue_fill = _fill("DDEBF7")   # Aprovado: nota >= 5

    for b in BIMESTRES:
        ws = wb.create_sheet(f"Notas - {b}")
        ws.sheet_properties.tabColor = "375623"

        headers = ["Nº", "Nome"] + DISCIPLINAS + ["Média Geral"]
        larguras = [5, 30, 14, 14, 12, 12, 12, 10, 13]
        for col, (h, w) in enumerate(zip(headers, larguras), 1):
            aplicar_header(ws.cell(row=1, column=col), h)
            ws.column_dimensions[get_column_letter(col)].width = w

        ws.row_dimensions[1].height = 25

        for row in range(2, N_ALUNOS + 2):
            ws.cell(row=row, column=1, value=row - 1)
            ws.cell(row=row, column=2).value = f"=Turma!B{row}"
            # Média das 6 disciplinas (cols C=3 a H=8)
            ws.cell(row=row, column=9).value = f'=IFERROR(AVERAGE(C{row}:H{row}),"")'
            ws.cell(row=row, column=9).number_format = "0.0"
            for col in range(1, 10):
                estilizar_dado(ws.cell(row=row, column=col))

        # Formatação condicional: vermelho (Reprovado) < 5, azul (Aprovado) >= 5
        grade_range = f"C2:I{N_ALUNOS + 1}"
        ws.conditional_formatting.add(
            grade_range, CellIsRule(operator="lessThan", formula=["5"], fill=red_fill)
        )
        ws.conditional_formatting.add(
            grade_range, CellIsRule(operator="greaterThanOrEqual", formula=["5"], fill=blue_fill)
        )

        ws.freeze_panes = "C2"

def add_recuperacao(wb):
    ws = wb.create_sheet("Recuperação")
    ws.sheet_properties.tabColor = "FF0000"

    headers = [
        ("Bimestre",       12),
        ("Nº Aluno",        9),
        ("Nome do Aluno",  30),
        ("Disciplina",     14),
        ("Nota Original",  14),
        ("Nota Recup.",    14),
        ("Situação Final", 16),
    ]
    for col, (titulo, largura) in enumerate(headers, 1):
        aplicar_header(ws.cell(row=1, column=col), titulo)
        ws.column_dimensions[get_column_letter(col)].width = largura

    ws.row_dimensions[1].height = 25

    # Pré-popular 40 linhas em branco com fórmula de Situação Final
    for row in range(2, 42):
        ws.cell(row=row, column=7).value = (
            f'=IF(F{row}="","",IF(F{row}>=5,"Aprovado","Em recuperação"))'
        )
        for col in range(1, 8):
            estilizar_dado(ws.cell(row=row, column=col))

    ws.conditional_formatting.add(
        "G2:G41",
        CellIsRule(operator="equal", formula=['"Aprovado"'], fill=_fill("DDEBF7"))
    )
    ws.conditional_formatting.add(
        "G2:G41",
        CellIsRule(operator="equal", formula=['"Em recuperação"'], fill=_fill("FFC7CE"))
    )

    ws.freeze_panes = "A2"


def add_contatos(wb):
    ws = wb.create_sheet("Contatos e Reuniões")
    ws.sheet_properties.tabColor = "FF9900"

    headers = [
        ("Data",          12),
        ("Nº Aluno",       9),
        ("Nome do Aluno", 30),
        ("Responsável",   25),
        ("Tipo",          14),
        ("Assunto",       40),
        ("Encaminhamento",40),
    ]
    for col, (titulo, largura) in enumerate(headers, 1):
        aplicar_header(ws.cell(row=1, column=col), titulo)
        ws.column_dimensions[get_column_letter(col)].width = largura

    ws.row_dimensions[1].height = 25

    for row in range(2, 102):
        for col in range(1, 8):
            c = ws.cell(row=row, column=col)
            c.border = _border()
            if col in [1, 2, 5]:
                c.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            else:
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = 40

    ws.freeze_panes = "A2"


def add_ocorrencias(wb):
    ws = wb.create_sheet("Ocorrências")
    ws.sheet_properties.tabColor = "9900FF"

    headers = [
        ("Data",           12),
        ("Nº Aluno",        9),
        ("Nome do Aluno",  30),
        ("Descrição",      50),
        ("Providência",    50),
    ]
    for col, (titulo, largura) in enumerate(headers, 1):
        aplicar_header(ws.cell(row=1, column=col), titulo)
        ws.column_dimensions[get_column_letter(col)].width = largura

    ws.row_dimensions[1].height = 25

    for row in range(2, 102):
        for col in range(1, 6):
            c = ws.cell(row=row, column=col)
            c.border = _border()
            c.alignment = Alignment(
                horizontal="center" if col in [1, 2] else "left",
                vertical="top",
                wrap_text=True
            )
        ws.row_dimensions[row].height = 40

    ws.freeze_panes = "A2"


def add_resumo(wb):
    ws = wb.create_sheet("Resumo Anual")
    ws.sheet_properties.tabColor = "404040"

    # Linha 1: título
    ws.merge_cells("A1:I1")
    t = ws.cell(row=1, column=1, value="RESUMO ANUAL — 1º ANO")
    t.font = Font(bold=True, size=14, color=COR_BRANCO)
    t.fill = _fill("404040")
    t.alignment = _center()
    ws.row_dimensions[1].height = 35

    # Linha 2: cabeçalhos
    headers = ["Nº", "Nome", "Freq. %"] + [f"Média {b}" for b in BIMESTRES] + ["Média Final", "Situação"]
    larguras = [5, 30, 9] + [11] * 4 + [12, 16]
    for col, (h, w) in enumerate(zip(headers, larguras), 1):
        aplicar_header(ws.cell(row=2, column=col), h)
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[2].height = 25

    for row in range(3, N_ALUNOS + 3):
        aluno_row = row - 1  # linha correspondente nas abas de notas (começa em 2)
        ws.cell(row=row, column=1, value=row - 2)
        ws.cell(row=row, column=2).value = f"=Turma!B{aluno_row}"

        # Frequência: média das % de todos os meses (cols AK nas abas de presença)
        freq_refs = "+".join([f"'Presença - {m}'!AK{aluno_row}" for m in MESES])
        ws.cell(row=row, column=3).value = f"=IFERROR(({freq_refs})/{len(MESES)},\"\")"
        ws.cell(row=row, column=3).number_format = "0.0%"

        # Médias por bimestre
        for b_idx, b in enumerate(BIMESTRES):
            col = 4 + b_idx
            ws.cell(row=row, column=col).value = f"='Notas - {b}'!I{aluno_row}"
            ws.cell(row=row, column=col).number_format = "0.0"

        # Média Final
        ws.cell(row=row, column=8).value = f"=IFERROR(AVERAGE(D{row}:G{row}),\"\")"
        ws.cell(row=row, column=8).number_format = "0.0"

        # Situação
        ws.cell(row=row, column=9).value = (
            f'=IF(H{row}="","",IF(AND(H{row}>=5,C{row}>=0.75),"Aprovado",'
            f'IF(H{row}>=5,"Rec. Freq.","Recuperação")))'
        )

        for col in range(1, 10):
            estilizar_dado(ws.cell(row=row, column=col))

    # Cores para Situação
    ws.conditional_formatting.add(
        f"I3:I{N_ALUNOS + 2}",
        CellIsRule(operator="equal", formula=['"Aprovado"'], fill=_fill("DDEBF7"))
    )
    ws.conditional_formatting.add(
        f"I3:I{N_ALUNOS + 2}",
        CellIsRule(operator="equal", formula=['"Recuperação"'], fill=_fill("FFC7CE"))
    )
    ws.conditional_formatting.add(
        f"I3:I{N_ALUNOS + 2}",
        CellIsRule(operator="equal", formula=['"Rec. Freq."'], fill=_fill("FFEB9C"))
    )

    ws.freeze_panes = "C3"

def main():
    saida = os.path.join(os.path.dirname(__file__), "..", "diario-de-classe.xlsx")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    add_turma(wb)
    add_presenca_sheets(wb)
    add_notas_sheets(wb)
    add_recuperacao(wb)
    add_contatos(wb)
    add_ocorrencias(wb)
    add_resumo(wb)

    wb.save(saida)
    print(f"Gerado: {os.path.abspath(saida)}")

if __name__ == "__main__":
    main()
